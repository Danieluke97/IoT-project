import requests
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
from datetime import datetime

class ThingSpeakSubscriber:
    def __init__(self, read_channel_id, read_api_key, write_channel_id, write_api_key):
        self.read_channel_id = read_channel_id
        self.read_api_key = read_api_key
        self.write_channel_id = write_channel_id
        self.write_api_key = write_api_key

    def download_data(self):
        url = f'https://api.thingspeak.com/channels/{self.read_channel_id}/feeds.csv?api_key={self.read_api_key}'
        response = requests.get(url)
        csv_file = 'thingSpeak_data.csv'
        with open(csv_file, 'wb') as file:
            file.write(response.content)
        print(f'Data exported to {csv_file}')
        return csv_file

    def normalize(self, values, min_val, max_val):
        if max_val == min_val:
            return [0] * len(values)
        return [(x - min_val) / (max_val - min_val) * 100 for x in values]

    def analyze_data(self, csv_file):
        data_df = pd.read_csv(csv_file)
        data_df.columns = ['created_at', 'entry_id', 'heart_rate', 'acceleration', 'gyro', 'stress', 'resp_rate', 'incline', 'set_number', 'exercise_name']

        # Convert 'created_at' to datetime and format as string
        data_df['created_at'] = pd.to_datetime(data_df['created_at']).dt.strftime('%Y-%m-%d %H:%M:%S')

        results = []
        
        min_max_values = {
            'heart_rate': (data_df['heart_rate'].min(), data_df['heart_rate'].max()),
            'acceleration': (data_df['acceleration'].min(), data_df['acceleration'].max()),
            'gyro': (data_df['gyro'].min(), data_df['gyro'].max()),
            'stress': (data_df['stress'].min(), data_df['stress'].max()),
            'resp_rate': (data_df['resp_rate'].min(), data_df['resp_rate'].max()),
            'incline': (data_df['incline'].min(), data_df['incline'].max())
        }

        grouped_df = data_df.groupby(['set_number', 'exercise_name'])

        for (set_number, exercise_name), group in grouped_df:
            heart_rate = group['heart_rate'].values
            accel = group['acceleration'].values
            gyro = group['gyro'].values
            stress = group['stress'].values
            resp_rate = group['resp_rate'].values
            incline = group['incline'].values
            created_at = group['created_at'].iloc[0]

            group = group.sort_values(by='created_at')
            time_diffs = (pd.to_datetime(group['created_at']).diff().fillna(pd.Timedelta(seconds=0)))
            time_under_threshold = time_diffs.sum().total_seconds()

            threshold = 1.5
            above_threshold = accel > threshold
            tut = time_under_threshold

            mass = 70
            velocity = np.cumsum(accel) / 50
            power = mass * accel * velocity
            total_power = np.sum(power)
            internal_load = np.mean(heart_rate) + np.mean(stress)
            external_load = total_power
            load_ratio = internal_load / external_load if external_load != 0 else 0

            efficiency = total_power / (np.mean(heart_rate) * np.mean(resp_rate)) if np.mean(heart_rate) * np.mean(resp_rate) != 0 else 0

            angle_variation = np.max(incline) - np.min(incline)
            max_gyro = np.max(gyro)
            normalized_angle_variation = self.normalize([angle_variation], min_max_values['incline'][0], min_max_values['incline'][1])[0]
            normalized_max_gyro = self.normalize([max_gyro], min_max_values['gyro'][0], min_max_values['gyro'][1])[0]
            risk_index = normalized_angle_variation * normalized_max_gyro

            resp_efficiency = total_power / np.mean(resp_rate) if np.mean(resp_rate) != 0 else 0
            rom = np.max(incline) - np.min(incline)
            specific_power = total_power / mass if mass != 0 else 0
            posture_variation = angle_variation

            results.append({
                'Date': created_at,
                'Set Number': set_number,
                'Exercise Name': exercise_name,
                'TUT': tut,
                'LoadRatio': load_ratio,
                'Efficiency': efficiency,
                'RiskIndex': risk_index,
                'RespEfficiency': resp_efficiency,
                'ROM': rom,
                'SpecificPower': specific_power,
                'PostureVariation': posture_variation
            })

        results_df = pd.DataFrame(results)

        # Normalize selected results
        for col in ['LoadRatio', 'Efficiency', 'RiskIndex', 'RespEfficiency', 'SpecificPower', 'PostureVariation']:
            min_val, max_val = results_df[col].min(), results_df[col].max()
            results_df[col] = self.normalize(results_df[col], min_val, max_val)

        return data_df, results_df

    def export_results(self):
        csv_file = self.download_data()
        data_df, results_df = self.analyze_data(csv_file)

        xlsx_file_name = 'gym_analysis_results.xlsx'
        with pd.ExcelWriter(xlsx_file_name, engine='openpyxl') as writer:
            data_df.to_excel(writer, sheet_name='OriginalData', index=False)
            results_df.to_excel(writer, sheet_name='AnalysisResults', index=False)

        workbook = load_workbook(xlsx_file_name)

        # Define date format style
        date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD HH:MM:SS')

        # Apply date format to 'created_at' column
        sheet = workbook['OriginalData']
        for cell in sheet['A'][1:]:
            cell.value = datetime.strptime(cell.value, '%Y-%m-%d %H:%M:%S')
            cell.style = date_style

        # Create a sheet for each exercise and add charts
        grouped_results = results_df.groupby('Exercise Name')
        
        for exercise_name, group in grouped_results:
            # Write data to the sheet
            if exercise_name in workbook.sheetnames:
                sheet = workbook[exercise_name]
            else:
                sheet = workbook.create_sheet(title=exercise_name)
            
            for r in dataframe_to_rows(group, index=False, header=True):
                sheet.append(r)
            
            # Create charts for this exercise
            charts_info = {
                'TUT': 'D',
                'LoadRatio': 'E',
                'Efficiency': 'F',
                'RiskIndex': 'G',
                'RespEfficiency': 'H',
                'ROM': 'I',
                'SpecificPower': 'J',
                'PostureVariation': 'K'
            }

            for idx, (chart_title, col) in enumerate(charts_info.items(), start=1):
                chart = LineChart()
                chart.title = chart_title
                chart.x_axis.title = "Set Number"
                chart.y_axis.title = chart_title

                data = Reference(sheet, min_col=column_index_from_string(col), min_row=2, max_row=sheet.max_row)
                cats = Reference(sheet, min_col=column_index_from_string('B'), min_row=2, max_row=sheet.max_row)
                chart.add_data(data, titles_from_data=False)
                chart.set_categories(cats)

                chart_location = f'{get_column_letter(14 + (idx - 1) * 10)}2'
                sheet.add_chart(chart, chart_location)

        # Save the workbook with the added charts
        workbook.save(xlsx_file_name)
        print(f'Data analyzed and saved in {xlsx_file_name} with separated charts for each exercise.')

if __name__ == "__main__":
    subscriber = ThingSpeakSubscriber(
        read_channel_id=2634877,
        read_api_key='PSR7PTZ3BH96VJFP',
        write_channel_id=2635570,
        write_api_key='J029IN9UIVOB8Q5F'
    )
    subscriber.export_results()
